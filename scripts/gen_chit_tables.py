"""One-off: derive the new Chit_Ledger + Chit_Taken_Member tables and fold them
into src/data/seed.json.

- Chit_Ledger  : one row per member per auction/month — their monthly due, what
  was received and what is still pending. This is the chit fund's OWN ledger,
  kept separate from the firm's Transaction_Ledger. (Supersedes Chit_Transaction.)
- Chit_Taken_Member : one row per auction winner — the payout owed to the member
  who took that month's chit, what has been given and what is still pending.

Reads the authoritative auction logic from the AppSheet export when present,
otherwise falls back to what is already in the seed.
"""
import io, json, os, glob, datetime

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SEED = os.path.join(ROOT, 'src', 'data', 'seed.json')

with io.open(SEED, encoding='utf-8') as f:
    data = json.load(f)

members = {m['Member_ID']: m for m in data.get('Chit_Member', [])}
auctions = {a['Chit_Auction_ID']: a for a in data.get('Chit_Auction', [])}


def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def member_name(mid):
    if mid in members:
        return members[mid]['Member_Name']
    if mid == 'Company_Chit':
        return 'Company Chit'
    # ID like Chit_A1_M17_Arun Prakash -> take the part after the 3rd underscore
    parts = str(mid).split('_', 3)
    return parts[3] if len(parts) > 3 else str(mid)


def ledger_status(due, recv, pend):
    if pend <= 0 and recv > 0:
        return 'Paid'
    if recv > 0:
        return 'Partial'
    return 'Pending'


# ── Chit_Ledger — derived from the existing per-member monthly Chit_Transaction ──
ledger = []
for t in data.get('Chit_Transaction', []):
    aid = t.get('Chit_Auction_ID')
    auc = auctions.get(aid, {})
    due = num(t.get('Indivitual_Member_Amount'))
    recv = num(t.get('Received_Amount'))
    pend = num(t.get('Pending_Amount')) if t.get('Pending_Amount') is not None else max(0.0, due - recv)
    status = ledger_status(due, recv, pend)
    ledger.append({
        'ID': t.get('ID'),
        'Finance_Name': auc.get('Finance_Name') or t.get('Finance_Name'),
        'Chit_ID': auc.get('Chit_ID'),
        'Chit_Name': t.get('Chit_Name'),
        'Chit_Auction_ID': aid,
        'Month_Count': t.get('Month_Count'),
        'Date_Auction': t.get('Date_Auction'),
        'Member_ID': t.get('Member_ID'),
        'Member_Name': t.get('Member_Name'),
        'Recommended_Partner': t.get('Recommended_Partner'),
        'Member_Percentage': t.get('Member_Percentage'),
        'One_Share_Amount': t.get('One_Share_Amount'),
        'Due_Amount': due,
        'Received_Amount': recv,
        'Pending_Amount': pend,
        'Payment_Type': 'Cash' if status != 'Pending' else None,
        'Paid_Date': t.get('Date_Auction') if status == 'Paid' else None,
        'Status': status,
    })


# ── Chit_Taken_Member — winners per auction, from the AppSheet export ───────────
def norm_date(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime('%Y-%m-%d')
    return v


taken = []
xlsx = None
cands = [c for c in glob.glob(os.path.join(os.path.expanduser('~'), 'Downloads', 'Finance_Details*.xlsx'))
         if not os.path.basename(c).startswith('~$')]
if cands:
    import openpyxl
    # Pick the export that actually has the Chit_Taken_Member sheet.
    for c in sorted(cands, key=len, reverse=True):
        try:
            if 'Chit_Taken_Member' in openpyxl.load_workbook(c, read_only=True).sheetnames:
                xlsx = c
                break
        except Exception:
            continue
if xlsx:
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb['Chit_Taken_Member']
    rows = list(ws.iter_rows(values_only=True))
    hdr = list(rows[0])
    idx = {h: i for i, h in enumerate(hdr)}

    def g(r, key):
        i = idx.get(key)
        return r[i] if i is not None and i < len(r) else None

    for r in rows[1:]:
        if not g(r, 'Chit_Taken_ID'):
            continue
        aid = g(r, 'Chit_Auction_ID')
        auc = auctions.get(aid, {})
        mid = g(r, 'Member_ID')
        total_to = num(g(r, 'Total_Amount_to_Member'))
        # The export stores #REF! for these; a taker's payout starts fully pending.
        given = num(members.get(mid, {}).get('Amount_Given'))
        pend = max(0.0, total_to - given)
        taken.append({
            'Chit_Taken_ID': g(r, 'Chit_Taken_ID'),
            'Chit_Auction_ID': aid,
            'Chit_ID': auc.get('Chit_ID') or g(r, 'Chit_ID'),
            'Chit_Name': g(r, 'Chit_Name'),
            'Date_Auction': norm_date(g(r, 'Date_Auction')),
            'Month_Count': g(r, 'Month_Count'),
            'Total_Auction_Amount': num(g(r, 'Total_Auction_Amount')),
            'Member_ID': mid,
            'Member_Name': member_name(mid),
            'Member_Type': g(r, 'Member_Type'),
            'Percentage_Need_to_Take': g(r, 'Percentage_Need_to_Take'),
            'Total_Amount_to_Member': total_to,
            'Amount_Given_to_Member': given,
            'Pending_Amount': pend,
            'Finance_Name': auc.get('Finance_Name') or g(r, 'Finance_Name'),
            'Need_to_Take_From_Previous_Company_Chit': g(r, 'Need_to_Take_From_Previous_Company_Chit'),
            'Amount_Taken_From_Company_Chit': num(g(r, 'Amount_Taken_From_Company_Chit')),
            'Remaining_Amount_in_Company_Chit': num(g(r, 'Remaining_Amount_in_Company_Chit')),
            'Status': 'Pending' if pend > 0 else 'Given',
        })
else:
    print('WARN: Excel not found, Chit_Taken_Member left as-is')
    taken = data.get('Chit_Taken_Member', [])

# ── Splice into seed.json, preserving key order; drop the old Chit_Transaction ──
out = {}
for k, v in data.items():
    if k == 'Chit_Transaction':
        out['Chit_Taken_Member'] = taken
        out['Chit_Ledger'] = ledger
    else:
        out[k] = v
if 'Chit_Ledger' not in out:      # seed had no Chit_Transaction key
    out['Chit_Taken_Member'] = taken
    out['Chit_Ledger'] = ledger

with io.open(SEED, 'w', encoding='utf-8') as f:
    json.dump(out, f, indent=0, ensure_ascii=False)

print(f'Chit_Ledger rows: {len(ledger)}  Chit_Taken_Member rows: {len(taken)}')
print('Wrote', SEED)
