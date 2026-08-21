"""Generate supabase/kannan_data.sql from Kannan_Finance.xlsx.

Imports the two finances in that workbook (New Finance, Kannnan_Personal) into the
existing Supabase tables. Columns + types are taken from src/data/seed.json so the
INSERTs match the tables migrate.sql already created. Idempotent: each table is
cleared for these two finances before inserting."""
import openpyxl
import json
import os
import datetime

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = r'C:\Users\ARUL\Downloads\Kannan_Finance.xlsx'
FINANCES = ['New Finance', 'Kannnan_Personal']

# Tables to import (all are Finance_Name-scoped in the app schema).
TABLES = [
    'Finance_Details', 'Partner', 'STL_CRM', 'Loan_Processing', 'Interest_Details',
    'Transaction_Ledger', 'Depositer_Interest', 'Other_Finance_Interest',
    'Deposit_Amount', 'Other_Finance_Loan', 'Chit_Creation', 'Chit_Member', 'Chit_Auction',
]

with open(os.path.join(ROOT, 'src', 'data', 'seed.json'), encoding='utf-8') as f:
    seed = json.load(f)


def col_type(vals):
    seen = False
    for v in vals:
        if v is None:
            continue
        seen = True
        if isinstance(v, bool) or not isinstance(v, (int, float)):
            return 'text'
    return 'numeric' if seen else 'text'


def schema(table):
    rows = seed.get(table, [])
    keys = []
    for r in rows:
        for k in r.keys():
            if k not in keys:
                keys.append(k)
    types = {k: col_type([r.get(k) for r in rows]) for k in keys}
    return keys, types


def clean(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.isoformat()[:10]
    return v


def sql_val(v, typ):
    v = clean(v)
    if v is None or v == '':
        return 'NULL'
    if typ == 'numeric':
        try:
            return str(int(v)) if isinstance(v, int) else str(float(v))
        except Exception:
            try:
                return str(float(str(v).replace(',', '')))
            except Exception:
                return 'NULL'
    return "'" + str(v).replace("'", "''") + "'"


wb = openpyxl.load_workbook(XLSX, data_only=True)
out = [
    '-- Kannan_Finance.xlsx import: New Finance + Kannnan_Personal',
    '-- Paste into Supabase > SQL Editor > Run. Safe to re-run (clears these two finances first).',
    '',
]
fin_list = ", ".join("'" + f.replace("'", "''") + "'" for f in FINANCES)
counts = {}

for table in TABLES:
    if table not in wb.sheetnames:
        continue
    ws = wb[table]
    hdr = [c.value for c in ws[1]]
    app_keys, app_types = schema(table)
    cols = [c for c in hdr if c in app_keys]  # intersection, in app order
    if 'Finance_Name' not in cols:
        continue
    idx = {c: hdr.index(c) for c in cols}
    fn_i = hdr.index('Finance_Name')

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[fn_i] not in FINANCES:
            continue
        rows.append([sql_val(r[idx[c]], app_types[c]) for c in cols])
    counts[table] = len(rows)
    if not rows:
        continue

    out.append(f'delete from "{table}" where "Finance_Name" in ({fin_list});')
    collist = ", ".join(f'"{c}"' for c in cols)
    for i in range(0, len(rows), 100):
        chunk = rows[i:i + 100]
        vals = ",\n".join('(' + ", ".join(row) + ')' for row in chunk)
        out.append(f'insert into "{table}" ({collist}) values\n{vals};')
    out.append('')

with open(os.path.join(ROOT, 'supabase', 'kannan_data.sql'), 'w', encoding='utf-8') as f:
    f.write("\n".join(out))
print('kannan_data.sql rows per table:', counts)
